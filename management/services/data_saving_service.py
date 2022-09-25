from management.services.answers import GetReplyWithValueChosenOper, GetReplyFilePath, GetReplyFileName
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from fpdf import FPDF
from datetime import date
from collections import namedtuple


class SavingService:
    def __init__(self):
        self._pdf_obj = OperationsPDF('P', 'mm', 'A4')

    def generate_statement(self, data: list, logged_in_user: namedtuple, used_account: namedtuple, dates: namedtuple):
        """Generate a statement for a specified period of time and save to the selected format"""
        if data:
            print("""
                Select the format in which you would like to save the data:
                1. XLSX
                2. PDF
            """)
            chosen_operation = GetReplyWithValueChosenOper().get_value(2)
            file_path = GetReplyFilePath().get_value()
            file_name = GetReplyFileName().get_value()
            match chosen_operation:
                case '1':
                    file_path = self.create_path(file_path, file_name, "xlsx")
                    self._save_data_to_xlsx(data, logged_in_user, file_path)
                case '2':
                    file_path = self.create_path(file_path, file_name, "pdf")
                    self._serialize_to_pdf(file_path, used_account, logged_in_user, data, dates)
        else:
            print(f"\n{' ' * 12}You don't have any history for the selected time period.")

    def _serialize_to_pdf(
            self, file_path: str, used_account: namedtuple, logged_in_user: namedtuple, data: list, dates: namedtuple):
        """Save data to pdf"""
        self._pdf_obj.statement_header = f'Statement {used_account.currency}'
        self._pdf_obj.title_name = f'{logged_in_user.name} {logged_in_user.surname}'
        self._pdf_obj.title_address = f'{logged_in_user.country}\n{logged_in_user.address}\n{logged_in_user.email}'
        self._pdf_obj.body_text = data
        self._pdf_obj.add_page()
        self._pdf_obj.print_user_data()
        self._pdf_obj.print_body(logged_in_user, used_account, dates)
        self._pdf_obj.output(file_path)

    @staticmethod
    def _save_data_to_xlsx(data: list, logged_in_user: namedtuple, file_path: str):
        wb = Workbook()
        ws = wb.active
        ws.append(["Data", "Customer", "Account number", "Card number", "Payout", "Payment", "Rate", "Saldo",
                   f"Fee in {logged_in_user.main_currency}"])
        [ws.append(row) for row in data]
        for letter in "ABCDEFGHI":
            ws.column_dimensions[letter].width = 23
        tab = Table(displayName="Table1", ref=f"A1:I{len(data) + 1}")
        style = TableStyleInfo(
            name="TableStyleMedium8", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(file_path)

    @staticmethod
    def create_path(file_path: str, file_name: str, format_: str) -> str:
        """Prepare a path with the filename"""
        return file_path + r'\\' + file_name + '.' + format_


class OperationsPDF(FPDF):
    def __init__(self, orientation: str, unit: str, format_: str, font_cache_dir="DEPRECATED"):
        super().__init__(orientation, unit, format_, font_cache_dir)
        self.banner_header = 'Currency Exchange App'
        self.statement_header = ""
        self.date_header = f'Generated {date.today()}'
        self.title_name = ''
        self.title_address = ''
        self.body_text = []

    def header(self):
        """Elements for the header in PDF"""
        self.set_font('helvetica', 'B', 16)
        title_w = self.get_string_width(self.banner_header) + 6
        self.set_x(10)
        self.ln(10)
        self.cell(title_w, 0, self.banner_header, border=0, ln=1, align='L', fill=False)
        self.cell(320, 0, self.statement_header, border=0, ln=1, align='C', fill=False)
        self.set_font('helvetica', '', 8)
        self.cell(320, 10, self.date_header, border=0, ln=1, align='C', fill=False)
        self.ln(20)

    def footer(self):
        """Elements for the footer in PDF"""
        self.set_y(-15)
        self.set_font('helvetica', 'I', 8)
        self.set_text_color(169, 169, 169)
        self.cell(0, 10, f'Page {self.page_no()}', align='C')

    def print_user_data(self):
        """Show user data in PDF"""
        title_name_w = self.get_string_width(self.banner_header) + 6
        title_address_w = self.get_string_width(self.banner_header) + 6
        self.set_x(10)
        self.set_font('helvetica', '', 14)
        self.cell(title_name_w, 5, self.title_name, border=0, ln=1, align='L', fill=False)
        self.set_font('helvetica', '', 8)
        self.multi_cell(title_address_w, 4, self.title_address, border=0, ln=1, align='L', fill=False)
        self.ln(20)

    def print_body(self, logged_in_user: namedtuple, used_account: namedtuple, dates: namedtuple):
        """Print the required text in the body section"""
        self.transaction_time(dates)
        self.get_statement_header()
        for row in self.body_text:
            if self.get_y() > 270:
                self.add_page('P', 'A4')
                self.get_statement_header()
            self.get_statement_first_line(used_account, row)
            self.get_statement_second_line(row, logged_in_user)
        self.ln()

    def transaction_time(self, dates: namedtuple):
        """Show information for which period of time the statement was made"""
        self.set_font('helvetica', 'B', 12)
        self.cell(10, 0, f"Account transactions from {str(dates.start_date.strftime('%d %B %Y'))}"
                         f" to {str(dates.end_date.strftime('%d %B %Y'))}", border=0, ln=1, align='L', fill=False)
        self.ln(10)

    def get_statement_header(self):
        """Show the header for the columns in which the data from the ttransaction will be displayed"""
        self.set_font('helvetica', 'B', 12)
        self.cell(10, 0, "Date", border=0, ln=1, align='C', fill=False)
        self.cell(140, 0, "Customer", border=0, ln=1, align='C', fill=False)
        self.cell(220, 0, "Payout", border=0, ln=1, align='C', fill=False)
        self.cell(300, 0, "Payment", border=0, ln=1, align='C', fill=False)
        self.cell(0, 0, "Balance", border=0, ln=1, align='R', fill=False)
        self.line(10, self.get_y()+3, 200, self.get_y()+3)
        self.ln(7)

    def get_statement_first_line(self, used_account: namedtuple, row: namedtuple):
        """Display main transaction data"""
        self.set_font('helvetica', 'I', 10)
        self.cell(50, 0, str(row.date), border=0, ln=1, align='L', fill=False)
        self.cell(140, 0, row.customer, border=0, ln=1, align='C', fill=False)
        self.cell(220, 0, self.currency(used_account.currency, row.payout), border=0, ln=1, align='C', fill=False)
        self.cell(300, 0, self.currency(used_account.currency, row.payment), border=0, ln=1, align='C', fill=False)
        self.cell(0, 0, self.currency(used_account.currency, row.saldo), border=0, ln=1, align='R', fill=False)
        self.ln(4)

    def get_statement_second_line(self, row: namedtuple, logged_in_user: namedtuple):
        """Display optional transaction data"""
        self.set_font('helvetica', 'I', 6)
        if row.acc_number and (not row.rate or row.rate == 1.00):
            self.cell(140, 0, f"Account nb: {row.acc_number}", border=0, ln=1, align='C', fill=False)
        elif row.acc_number:
            self.cell(140, 0, f"Account nb: {row.acc_number} Rate: {row.rate}", border=0, ln=1, align='C', fill=False)
        elif row.card_nb == "NOT EXIST" and (not row.rate or row.rate == 1.00):
            self.cell(140, 0, "", border=0, ln=1, align='C', fill=False)
        elif row.card_nb == "NOT EXIST":
            self.cell(140, 0, f"Rate: {row.rate}", border=0, ln=1, align='C', fill=False)
        elif row.card_nb and (not row.rate or row.rate == 1.00):
            self.cell(140, 0, f"Card nb: {row.card_nb}", border=0, ln=1, align='C', fill=False)
        elif row.card_nb:
            self.cell(140, 0, f"Card nb: {row.card_nb} Rate: {row.rate}", border=0, ln=1, align='C', fill=False)
        if row.commission and row.payout:
            self.cell(220, 0, f"Fee: {self.currency(logged_in_user.main_currency, row.commission)}",
                      border=0, ln=1, align='C', fill=False)
        self.line(10, self.get_y() + 3, 200, self.get_y() + 3)
        self.ln(6)

    @staticmethod
    def currency(currency: str, row: str) -> str:
        """Combine the currency symbol with the transaction amount"""
        if row:
            match currency:
                case "GBP":
                    return "£" + str(row)
                case "USD":
                    return "$" + str(row)
                case "EUR":
                    return "EUR " + str(row)
                case "CHF":
                    return "CHF " + str(row)
        return ""
